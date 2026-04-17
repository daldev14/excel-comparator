from pathlib import Path
from openpyxl import load_workbook


def compare_workbooks(workbook_path1: str, workbook_path2: str):

    workbook_file1 = load_workbook(Path(workbook_path1), data_only=True)
    workbook_file2 = load_workbook(Path(workbook_path2), data_only=True)

    wb_sheets1 = set[str](workbook_file1.sheetnames)
    wb_sheets2 = set[str](workbook_file2.sheetnames)

    common_sheets = sorted(wb_sheets1 & wb_sheets2)
    missing_in_wbfile1 = sorted(wb_sheets2 - wb_sheets1)
    missing_in_wbfile2 = sorted(wb_sheets1 - wb_sheets2)

    differences = []

    for sheet_name in common_sheets:
        wbsheet1 = workbook_file1[sheet_name]
        wbsheet2 = workbook_file2[sheet_name]

        max_row = max(wbsheet1.max_row, wbsheet2.max_row)
        max_column = max(wbsheet1.max_column, wbsheet2.max_column)

        for row_idx in range(1, max_row + 1):
            for col_idx in range(1, max_column + 1):
                wbvalue1 = wbsheet1.cell(row=row_idx, column=col_idx).value
                wbvalue2 = wbsheet2.cell(row=row_idx, column=col_idx).value

                if wbvalue1 != wbvalue2:
                    differences.append(
                        {
                            "sheet_name": sheet_name,
                            "row": row_idx,
                            "column": col_idx,
                            "workbook_file1": wbvalue1,
                            "workbook_file2": wbvalue2,
                        }
                    )

    if not differences:
        print("No se encontraron diferencias.")
        return None

    if missing_in_wbfile1:
        print(f"Hojas faltantes en archivo 1: {', '.join(missing_in_wbfile1)}")

    if missing_in_wbfile2:
        print(f"Hojas faltantes en archivo 2: {', '.join(missing_in_wbfile2)}")

    print(f"Diferencias de celdas detectadas: {len(differences)}")
    for diff in differences:
        print(
            f"[{diff['sheet_name']}] R{diff['row']}C{diff['column']}: "
            f"{diff['workbook_file1']!r} != {diff['workbook_file2']!r}"
        )

    return differences
